from flask import Flask, render_template, request, jsonify, send_from_directory, send_file
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user
from peewee import *
from datetime import datetime as dt
import datetime
import os
from werkzeug.utils import secure_filename
import unicodedata
import re
from openpyxl import Workbook
import threading
import telebot


#-------- Chaves e Tokens a serem salvos no ambiente virtual

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'static/img'
UPLOAD_FOLDER_DB = os.path.dirname(os.path.abspath(__file__))
app.config['UPLOAD_FOLDER_DB'] = UPLOAD_FOLDER_DB
app.secret_key = os.getenv('SECRET_KEY')
app.config['usuario'] = os.getenv('DATABASE_USUARIO')
app.config['senha'] = os.getenv('DATABASE_SENHA')
TOKEN = os.getenv('TELEGRAM_TOKEN')

db = PostgresqlDatabase(
    'railway',  # Nome do banco
    user= os.getenv('POSTGRES_USUARIO'),
    password= os.getenv('POSTGRES_SENHA'),
    host= os.getenv('POSTGRES_HOST'),
    port= os.getenv('POSTGRES_PORTA')
)

#------------TELEGRAM CONFIGS --------------------------------------
# Configuração do Telegram Bot
bot = telebot.TeleBot(TOKEN)


@bot.message_handler(commands=['start'])
def send_welcome(message):
    # Busca todos os nomes na tabela Pessoal
    nomes = Pessoal.select().order_by(Pessoal.nome)

    if nomes.exists():
        # Cria os botões inline com os nomes e postos
        markup = telebot.types.InlineKeyboardMarkup()
        for pessoa in nomes:
            # Exibe nome e posto no botão
            markup.add(
                telebot.types.InlineKeyboardButton(
                    f"{pessoa.posto} {pessoa.nome}",  # Exibe "Posto Nome"
                    callback_data=f"pessoa_{pessoa.id}"
                )
            )

        bot.send_message(message.chat.id, "Pra prosseguir com o cadastro, selecione seu nome:", reply_markup=markup)
    else:
        bot.send_message(message.chat.id, "Não há nomes cadastrados no momento.")


@bot.callback_query_handler(func=lambda call: call.data.startswith('pessoa_'))
def handle_nome_selection(call):
    # Extrai o ID da pessoa a partir do callback_data
    pessoa_id = int(call.data.split('_')[1])
    pessoa = Pessoal.get_by_id(pessoa_id)

    # Captura o chat_id do usuário do Telegram
    chat_id = call.message.chat.id

    # Atualiza o campo chat_id no banco de dados
    pessoa.chat_id = chat_id
    print(chat_id)
    pessoa.save()

    # Responde a interação com uma mensagem de confirmação
    bot.answer_callback_query(call.id, f"Você selecionou: {pessoa.posto} {pessoa.nome}")
    bot.send_message(call.message.chat.id, f"Olá, {pessoa.posto} {pessoa.nome}! Seu chat_id foi registrado com sucesso. \n Para sua segurança, você receberá notificações quando realizar compras em seu nome!")

# Função para rodar o Flask
def run_flask():
    app.run(host='0.0.0.0', port=5000)

# Função para rodar o Telebot
def run_telebot():
    bot.infinity_polling()
#-------------LOGIN MANAGER ---------------------------------------
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'pagina_login'

#-------------CLASSES BANCO DE DADOS ------------------------------

class Pessoal(Model):
    id = AutoField(primary_key=True)
    posto = CharField()
    nome = CharField(unique=True)
    telefone = BigIntegerField(null=True)
    chat_id = BigIntegerField(null=True)

    class Meta:
        database = db
        db_table = 'pessoal'

# Modelo da tabela "estoque"
class Estoque(Model):
    id = AutoField(primary_key=True)
    produto = CharField(unique=True)
    tamanho = IntegerField(null=True)
    categoria = CharField()
    unidade = CharField()
    quantidade = IntegerField(null=True)
    preco_venda = FloatField()
    preco_compra = FloatField(null=True)
    foto = CharField(null=True)

    class Meta:
        database = db
        db_table = 'estoque'


class Venda(Model):
    id = AutoField(primary_key=True)
    nome = ForeignKeyField(Pessoal, field='nome', on_update='CASCADE', on_delete='CASCADE')
    produto = ForeignKeyField(Estoque, field='produto', on_update='CASCADE', on_delete='CASCADE')
    quantidade = IntegerField()
    valor_total = FloatField()
    hora = DateTimeField(default=datetime.datetime.now)
    data = DateField(default=datetime.date.today)

    class Meta:
        database = db
        db_table = 'venda'

class Compra(Model):
    id = AutoField(primary_key=True)
    produto = ForeignKeyField(Estoque, field='produto', on_update='CASCADE', on_delete='CASCADE')
    preco_compra = FloatField()
    quantidade = IntegerField()
    data = DateField()

    class Meta:
        database = db
        db_table = 'compra'


class User(UserMixin):
    def __init__(self, id):
        self.id = id


# Conexão e criação de tabelas
try:
    db.connect(reuse_if_open=True)
    db.create_tables([Pessoal, Estoque, Venda, Compra])
    print("Conectado ao banco e tabelas criadas!")
except Exception as e:
    print(f"Erro ao conectar ao banco: {e}")
finally:
    db.close()
#--------------------- FIM CLASSES BANCO DE DADOS --------------------



# ---------------------RENDERIZAÇÃO DE PÁGINAS -----------------------
@app.route("/")
def index():
    return render_template('index.html')

@login_manager.user_loader
def load_user(user_id):
    return User(user_id) if user_id == app.config['usuario'] else None

@app.route("/pagina_login")
def pagina_login():
    return render_template('login.html')

@app.route("/protected")
@login_required
def protected():
    return render_template('admin.html')

@app.route("/sucesso")
def sucesso():
    return render_template('sucesso.html')

@app.route("/error")
def error():
    return render_template('error.html')
# ---------------------FIM DE RENDERIZAÇÃO DE PÁGINAS -----------------------
@app.route("/login", methods=['POST'])
def login():
    data = request.json
    usuario = data.get('usuario')
    senha = data.get('senha')

    if usuario == app.config['usuario'] and senha == app.config['senha']:
        user = User(id=usuario)
        login_user(user)
        return jsonify({"status": "success", "message": "Usuário logado com sucesso"})

    return jsonify({"status": "error", "message": "Credenciais inválidas"}), 401

@app.route("/logout", methods=['POST'])
@login_required
def logout():
    logout_user()
    return jsonify({"status": "success", "message": "Usuário deslogado com sucesso"})

# ---------------------FUNÇÕES MANIPULAÇÃO DE BANCO DE DADOS -----------------------
@app.route('/cadastrar_pessoal', methods=['POST'])
def cadastrar_pessoal():
    try:
        data = request.json  # Captura os dados enviados
        print("Dados recebidos:", data)

        # Validação básica para evitar campos ausentes
        if not data:
            return jsonify({"status": "error", "message": "Nenhum dado recebido"}), 400

        posto = data.get('posto')
        nome = data.get('nome')
        telefone = data.get('telefone')
        chat_id = data.get('chat_id')

        print(f"Dados processados: posto={posto}, nome={nome}, telefone={telefone}, chat_id={chat_id}")

        # Criação do registro no banco
        Pessoal.create(posto=posto, nome=nome, telefone=telefone, chat_id=chat_id)

        return jsonify({"status": "success", "message": "Militar cadastrado com sucesso!"}), 201

    except IntegrityError as e:
        print("Erro de integridade:", e)
        return jsonify({"status": "error", "message": "Militar já cadastrado no banco de dados"}), 400

    except Exception as e:
        print("Erro inesperado:", e)
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/mostrar_pessoal')
def mostrar_pessoal():
    pessoal = Pessoal.select().order_by(Pessoal.nome)
    lista_pessoal =[]

    for pessoais in pessoal:
        lista_pessoal.append({
            'id': pessoais.id,
            'posto': pessoais.posto,
            'nome': pessoais.nome,
            'telefone': pessoais.telefone,
            'chat_id' : pessoais.chat_id
        })
    return jsonify(lista_pessoal)

@app.route('/remover_pessoal/<int:id>', methods=['DELETE'])
def remover_pessoal(id):
    try:
        pessoal = Pessoal.get_or_none(Pessoal.id == id)

        if pessoal:
            pessoal.delete_instance()
            return jsonify({'message': 'Militar removido com sucesso!'}), 200
        else:
            return jsonify({'error': 'Militar não encontrada!'}), 404

    except Exception as e:
        return jsonify({'error': str(e)}), 500




@app.route('/cadastrar_produto', methods=['POST'])
def cadastrar_produto():
    def tratar_string(s):
        # Remove acentos
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        # Remove espaços
        s = re.sub(r'[^a-zA-Z0-9]', '', s)  # Remove tudo que não for letra ou número
        return s

    produto = request.form.get('produto')
    tamanho = request.form.get('tamanho')
    unidade = request.form.get('unidade')
    quantidade = 0
    categoria = request.form.get('categoria')
    preco_venda = float(request.form.get('preco_venda'))
    preco_compra = 0
    foto = f'static/img/{tratar_string(produto)}.png' # o que vai ser salvo no banco de dados
    foto_salva = request.files.get('foto') # a que vai ser salva na pasta /static/img

    # Validações básicas
    if not all([produto, tamanho, unidade, categoria, preco_venda, foto_salva]):
        return jsonify({'status': 'error', 'message': 'Todos os campos são obrigatórios'}), 400

    # Salva no banco de dados
    try:
        Estoque.create(produto=produto, tamanho=tamanho, unidade=unidade, categoria=categoria, preco_venda=preco_venda, foto=foto, quantidade=quantidade, preco_compra=preco_compra)
    except IntegrityError as e:
        return jsonify({'status': 'error', 'message': f'Parece que este produto já está cadastrado com esse nome: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Erro ao salvar no banco de dados: {str(e)}'}), 500

    # Processa o arquivo de imagem
    try:
        filename = f"{secure_filename(tratar_string(produto))}.png"
        foto_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        foto_salva.save(foto_path)
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Erro ao salvar imagem: {str(e)}'}), 500

    return jsonify({'status': 'success', 'message': 'Produto cadastrado com sucesso!'})

@app.route('/mostrar_produtos')
def mostrar_produtos():
    estoque = Estoque.select().order_by(Estoque.categoria.asc(), Estoque.produto.asc())
    lista_estoque =[]

    for produto in estoque:
        lista_estoque.append({
            'id': produto.id,
            'foto': produto.foto,
            'produto': produto.produto,
            'tamanho': produto.tamanho,
            'unidade': produto.unidade,
            'quantidade': produto.quantidade,
            'categoria': produto.categoria,
            'preco_venda': produto.preco_venda,
            'preco_compra': produto.preco_compra
        })
    return jsonify(lista_estoque)

@app.route('/remover_produto/<int:id>', methods=['DELETE'])
def remover_produto(id):
    try:
        produto = Estoque.get_or_none(Estoque.id == id)

        if produto:
            produto.delete_instance()
            return jsonify({'message': 'Produto removido com sucesso!'}), 200
        else:
            return jsonify({'error': 'Produto não encontrada!'}), 404

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/alterar_estoque/<int:id>', methods=['POST'])
def alterar_estoque(id):
    data = request.json
    preco_venda = data.get('preco_venda')  # Correção aqui

    try:
        produto = Estoque.get_or_none(Estoque.id == id)
        if produto is None:
            return jsonify({'error': 'Produto não encontrado!'}), 404

        produto.preco_venda = float(preco_venda)
        produto.save()

        return jsonify({'message': 'Produto alterado com sucesso!'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/cadastrar_compra', methods=['POST'])
def cadastrar_compra():
    data = request.json
    produto_nome = data.get('produto')  # Nome do produto ao invés do ID
    preco_compra = data.get('preco_compra')
    quantidade = data.get('quantidade')
    data_compra = data.get('data')


    try:
        # Verifica se o produto existe no estoque usando o nome
        produto = Estoque.get(Estoque.produto == produto_nome)

        # Cria o registro de compra
        Compra.create(produto=produto, preco_compra=preco_compra, quantidade=quantidade, data=data_compra)

        # Atualiza a tabela Estoque
        produto.quantidade += int(quantidade)
        produto.preco_compra = float(preco_compra)
        produto.save()

        return jsonify({"status": "success"}), 200

    except Estoque.DoesNotExist:
        return jsonify({"status": "error", "message": "Produto não encontrado no estoque."}), 404
    except IntegrityError as e:
        return jsonify(
            {"status": "error", "message": "Erro ao cadastrar compra. Produto já existente no banco de dados"}), 400


@app.route('/mostrar_compras')
def mostrar_compras():
    compras = Compra.select().order_by(Compra.data.desc())
    lista_compras =[]

    for compra in compras:
        lista_compras.append({
            'id': compra.id,
            'produto': compra.produto.produto,
            'quantidade': compra.quantidade,
            'preco_compra': compra.preco_compra,
            'data': compra.data,

        })
    return jsonify(lista_compras)

@app.route('/remover_compra/<int:id>', methods=['DELETE'])
def remover_compra(id):
    try:
        compra = Compra.get_or_none(Compra.id == id)#acessar dados de compras
        produto = Estoque.get(Estoque.id == compra.produto.id)#acessar dados do estoque

        produto.quantidade -= int(compra.quantidade)
        produto.save()

        if compra:
            compra.delete_instance()
            return jsonify({'message': 'Compra removida com sucesso!'}), 200
        else:
            return jsonify({'error': 'Compra não encontrada!'}), 404

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/cadastrar_venda', methods=['POST'])
def cadastrar_venda():
    data = request.json
    print("Dados recebidos:", data)

    nome = data.get('nome')
    produtosSelecionados = data.get('produtosSelecionados')

    # Validações iniciais
    if not nome:
        return jsonify({'status': 'error', 'message': 'Nome do cliente é obrigatório'}), 400

    if not produtosSelecionados or not isinstance(produtosSelecionados, list):
        return jsonify({'status': 'error', 'message': 'Lista de produtos inválida ou vazia'}), 400

    pessoa = Pessoal.get_or_none(Pessoal.nome == nome)

    if not pessoa:
        return jsonify({'status': 'error', 'message': 'Pessoa não encontrada'}), 404

    try:
        for venda in produtosSelecionados:
            # Valida cada produto
            if not isinstance(venda, dict):
                return jsonify({'status': 'error', 'message': 'Formato de produto inválido'}), 400

            produto_nome = venda.get('produto')
            quantidade = venda.get('quantidade')
            preco_total = venda.get('preco_total')

            if not produto_nome or not quantidade or not preco_total:
                return jsonify({'status': 'error', 'message': f'Dados incompletos para o produto: {venda}'}), 400

            quantidade = int(quantidade)
            preco_total = float(preco_total)

            # Busca o produto no estoque
            produto = Estoque.get_or_none(Estoque.produto == produto_nome)
            if not produto:
                return jsonify({'status': 'error', 'message': f'Produto "{produto_nome}" não encontrado'}), 404

            # Cria a venda
            Venda.create(
                nome=pessoa,
                produto=produto,
                quantidade=quantidade,
                valor_total=preco_total
            )


            produto.quantidade -= quantidade
            produto.save()


    except IntegrityError as e:
        print("Erro de integridade:", e)
        return jsonify({"status": "error", "message": "Erro na venda"}), 400
    except Exception as e:
        print("Erro inesperado ao salvar a venda:", e)
        return jsonify({'status': 'error', 'message': f'Erro interno ao processar a venda: {str(e)}'}), 500

    # Envia mensagem no Telegram se o chat_id estiver registrado
    if pessoa.chat_id:
        try:
            produtos = ', '.join([str(venda.get('produto', 'Desconhecido')) for venda in produtosSelecionados])
            total = sum(float(venda.get('preco_total', 0)) for venda in produtosSelecionados)
            bot.send_message(
                pessoa.chat_id,
                f"Olá {nome}, \nCompra realizada com sucesso! \nProdutos: {produtos}.\nTotal: R$ {total:.2f}. \n\nSe não foi você quem realizou essa compra, contate a administração do Grêmio!"
            )
        except Exception as e:
            print(f"Erro ao enviar mensagem no Telegram: {e}")

    print("Venda cadastrada com sucesso.")
    return jsonify({'status': 'success', 'message': 'Venda cadastrada com sucesso!'})



@app.route('/mostrar_vendas')
def mostrar_vendas():
    vendas = Venda.select()
    lista_venda =[]

    for venda in vendas:
        lista_venda.append({
            'id': venda.id,
            'nome': venda.nome.nome,
            'produto': venda.produto.produto,
            'quantidade': venda.quantidade,
            'valor_total': venda.valor_total,
            'hora': venda.hora,
            'data': venda.data
        })
    return jsonify(lista_venda)

@app.route('/remover_venda/<int:id>', methods=['DELETE'])
def remover_venda(id):
    try:
        venda = Venda.get_or_none(Venda.id == id)#acessar dados da venda
        produto = Estoque.get(Estoque.produto == venda.produto.produto)#acessar dados do estoque

        produto.quantidade += int(venda.quantidade)
        produto.save()

        if venda:
            venda.delete_instance()
            return jsonify({'message': 'Venda removida com sucesso!'}), 200
        else:
            return jsonify({'error': 'Venda não encontrada!'}), 404

    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/download_relatorio', methods=['POST'])
@login_required
def download_relatorio():
    data = request.json
    data1 = data.get('data1')
    data2 = data.get('data2')
    data_inicio = dt.strptime(data1, "%Y-%m-%d")
    data_fim = dt.strptime(data2, "%Y-%m-%d")

    # Filtra as vendas no intervalo de datas
    relatorio = Venda.select().where((Venda.data >= data_inicio) & (Venda.data <= data_fim))

    # Organiza os dados em dicionários
    dados = {}
    produtos = set()

    for venda in relatorio:
        nome = venda.nome.nome
        produto_nome = venda.produto.produto
        quantidade = venda.quantidade
        valor_total = venda.valor_total

        # Adiciona o produto ao conjunto de produtos
        produtos.add(produto_nome)

        # Organiza dados por pessoa
        if nome not in dados:
            dados[nome] = {'produtos': {}, 'total': 0}

        # Atualiza a quantidade e o valor total
        dados[nome]['produtos'][produto_nome] = dados[nome]['produtos'].get(produto_nome, 0) + quantidade
        dados[nome]['total'] += valor_total


    # Ordena os produtos para a ordem das colunas no Excel
    produtos = sorted(produtos)

    # Cria o arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Vendas"

    # Escreve o cabeçalho
    ws.cell(row=1, column=1, value="Nome")
    for col_index, produto in enumerate(produtos, start=2):
        ws.cell(row=1, column=col_index, value=produto)
    ws.cell(row=1, column=len(produtos) + 2, value="Total Consumido")

    # Preenche os dados por pessoa e produtos
    for row_index, (nome, info) in enumerate(dados.items(), start=2):
        ws.cell(row=row_index, column=1, value=nome)

        for col_index, produto in enumerate(produtos, start=2):
            quantidade = info['produtos'].get(produto, 0)
            ws.cell(row=row_index, column=col_index, value=quantidade)

        ws.cell(row=row_index, column=len(produtos) + 2, value=info['total'])

    # Salva o arquivo temporário
    file_name = f"relatorio_vendas - {data_inicio.strftime('%Y-%m-%d')} a {data_fim.strftime('%Y-%m-%d')}.xlsx"
    file_path = os.path.join('relatorios', file_name)
    wb.save(file_path)

    # Retorna o arquivo como resposta
    return send_file(file_path, as_attachment=True, download_name=file_name)




@app.route('/download_database')
@login_required
def download_database():
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER_DB'], 'dados.db', as_attachment=True)
    except Exception as e:
        print("Erro ao enviar o arquivo:", e)
        return {"message": "Erro ao enviar o arquivo."}, 500




# ---------------------FIM FUNÇÕES MANIPULAÇÃO DE BANCO DE DADOS -----------------------


if __name__ == "__main__":
    # Criar e iniciar as threads
    flask_thread = threading.Thread(target=run_flask)
    telebot_thread = threading.Thread(target=run_telebot)

    flask_thread.start()
    telebot_thread.start()

    flask_thread.join()
    telebot_thread.join()


